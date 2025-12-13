#!/usr/bin/env python3
"""Extract structural metadata from an Excel file for auditing."""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import zipfile
import xml.etree.ElementTree as ET

# Try to import xlrd for XLS support
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


def extract_structure_xls(filepath: str) -> dict:
    """Extract structure from XLS (Excel 97-2003) files using xlrd."""
    result = {
        "filename": Path(filepath).name,
        "format": "xls",
        "support_level": "basic",
        "sheets": [],
        "named_ranges": [],
        "tables": [],  # XLS doesn't have formal tables
        "external_links": [],
        "has_vba": False,
        "has_hidden_sheets": False,
        "file_size_mb": round(Path(filepath).stat().st_size / (1024 * 1024), 2),
        "summary": {}
    }

    try:
        # xlrd 2.0+ only supports xls, not xlsx
        wb = xlrd.open_workbook(filepath, on_demand=True)
    except Exception as e:
        result["error"] = f"Failed to load XLS: {str(e)}"
        return result

    # Extract sheet info
    for idx, sheet_name in enumerate(wb.sheet_names()):
        try:
            ws = wb.sheet_by_index(idx)

            # Determine visibility
            visibility = "visible"
            if hasattr(wb, 'sheet_visibility'):
                vis = wb.sheet_visibility(idx)
                if vis == 1:
                    visibility = "hidden"
                elif vis == 2:
                    visibility = "very_hidden"

            sheet_info = {
                "name": sheet_name,
                "visibility": visibility,
                "dimensions": f"A1:{xlrd.colname(ws.ncols-1) if ws.ncols > 0 else 'A'}{ws.nrows}" if ws.nrows > 0 else "empty",
                "row_count": ws.nrows,
                "col_count": ws.ncols,
                "has_filters": False,  # Can't easily detect in xlrd
                "freeze_panes": None,  # Limited support in xlrd
                "merged_cells_count": len(ws.merged_cells) if hasattr(ws, 'merged_cells') else 0,
            }

            # Sample headers (first row)
            headers = []
            if ws.nrows > 0:
                for col in range(min(ws.ncols, 25)):
                    try:
                        cell_value = ws.cell_value(0, col)
                        if cell_value:
                            headers.append(str(cell_value)[:50])
                    except:
                        pass
            sheet_info["headers_sample"] = headers

            result["sheets"].append(sheet_info)

            if visibility != "visible":
                result["has_hidden_sheets"] = True

        except Exception as e:
            result["sheets"].append({
                "name": sheet_name,
                "error": str(e)[:100]
            })

    # Extract named ranges
    try:
        for i in range(wb.name_obj_list.__len__() if hasattr(wb, 'name_obj_list') else 0):
            try:
                name_obj = wb.name_obj_list[i]
                result["named_ranges"].append({
                    "name": name_obj.name,
                    "refers_to": name_obj.formula_text[:100] if hasattr(name_obj, 'formula_text') else None,
                    "scope": "global" if name_obj.scope == -1 else f"sheet_{name_obj.scope}"
                })
            except:
                pass
    except:
        pass

    # Check for external links in formulas
    external_patterns = set()
    for idx in range(len(wb.sheet_names())):
        try:
            ws = wb.sheet_by_index(idx)
            for row in range(min(ws.nrows, 500)):
                for col in range(ws.ncols):
                    try:
                        cell = ws.cell(row, col)
                        # xlrd cell type 2 is formula (but we get the value, not formula text usually)
                        # For external refs, check if there's a formula with [
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            val = str(cell.value)
                            if val.startswith('=') and '[' in val and ']' in val:
                                start = val.find('[')
                                end = val.find(']')
                                if start < end:
                                    external_patterns.add(val[start+1:end])
                    except:
                        pass
        except:
            pass

    result["external_links"] = list(external_patterns)[:20]

    # Check for VBA - xlrd doesn't directly expose this, but we can check file structure
    try:
        with open(filepath, 'rb') as f:
            header = f.read(8)
            # OLE compound document with potential VBA
            if header[:4] == b'\xd0\xcf\x11\xe0':
                # Could have VBA, but can't confirm without deeper parsing
                # Mark as potential
                result["has_vba"] = False  # Conservative - can't confirm
    except:
        pass

    # Summary stats
    result["summary"] = {
        "total_sheets": len(result["sheets"]),
        "visible_sheets": sum(1 for s in result["sheets"] if s.get("visibility") == "visible"),
        "hidden_sheets": sum(1 for s in result["sheets"] if s.get("visibility") != "visible"),
        "total_named_ranges": len(result["named_ranges"]),
        "total_tables": 0,  # XLS doesn't have formal tables
        "external_link_count": len(result["external_links"]),
        "has_vba": result["has_vba"],
        "risk_flags": [],
        "format_note": "XLS format (Excel 97-2003) - basic structure extraction only"
    }

    # Add risk flags
    if result["has_hidden_sheets"]:
        result["summary"]["risk_flags"].append("Has hidden sheets")
    if len(result["external_links"]) > 0:
        result["summary"]["risk_flags"].append(f"External links: {len(result['external_links'])}")
    if any(s.get("merged_cells_count", 0) > 50 for s in result["sheets"]):
        result["summary"]["risk_flags"].append("Heavy use of merged cells")

    wb.release_resources()
    return result


def extract_structure_xlsx(filepath: str) -> dict:
    """Extract comprehensive structure from XLSX (Excel 2007+) files using openpyxl."""
    result = {
        "filename": Path(filepath).name,
        "format": "xlsx",
        "support_level": "full",
        "sheets": [],
        "named_ranges": [],
        "tables": [],
        "external_links": [],
        "has_vba": False,
        "has_hidden_sheets": False,
        "file_size_mb": round(Path(filepath).stat().st_size / (1024 * 1024), 2),
        "summary": {}
    }

    # Check for VBA (macros) by inspecting the zip
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            result["has_vba"] = any('vbaProject' in name for name in z.namelist())
    except:
        pass

    # Load workbook
    try:
        wb = load_workbook(filepath, read_only=False, data_only=False)
    except Exception as e:
        result["error"] = f"Failed to load: {str(e)}"
        return result

    # Extract sheet info
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "visibility": "visible" if ws.sheet_state == "visible" else ws.sheet_state,
            "dimensions": ws.dimensions or "empty",
            "row_count": ws.max_row or 0,
            "col_count": ws.max_column or 0,
            "has_filters": ws.auto_filter.ref is not None if ws.auto_filter else False,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "merged_cells_count": len(ws.merged_cells.ranges),
        }

        # Sample headers (first row)
        headers = []
        if ws.max_row and ws.max_row > 0:
            for col in range(1, min(ws.max_column + 1, 26)):  # First 25 cols
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    headers.append(str(cell.value)[:50])
        sheet_info["headers_sample"] = headers

        result["sheets"].append(sheet_info)

        if sheet_info["visibility"] != "visible":
            result["has_hidden_sheets"] = True

    # Extract named ranges
    for name in wb.defined_names:
        try:
            defn = wb.defined_names[name]
            result["named_ranges"].append({
                "name": name,
                "refers_to": str(defn.attr_text)[:100] if hasattr(defn, 'attr_text') else str(defn.value)[:100] if hasattr(defn, 'value') else None,
                "scope": "global"
            })
        except:
            pass

    # Extract tables
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for table in ws.tables.values():
            result["tables"].append({
                "name": table.name,
                "sheet": sheet_name,
                "range": table.ref,
                "display_name": table.displayName
            })

    # Check for external links in formulas (basic check)
    external_patterns = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=min(ws.max_row or 0, 500)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    if '[' in formula and ']' in formula:
                        # Extract external reference
                        start = formula.find('[')
                        end = formula.find(']')
                        if start < end:
                            ext_ref = formula[start+1:end]
                            external_patterns.add(ext_ref)

    result["external_links"] = list(external_patterns)[:20]  # Cap at 20

    # Summary stats
    result["summary"] = {
        "total_sheets": len(result["sheets"]),
        "visible_sheets": sum(1 for s in result["sheets"] if s["visibility"] == "visible"),
        "hidden_sheets": sum(1 for s in result["sheets"] if s["visibility"] != "visible"),
        "total_named_ranges": len(result["named_ranges"]),
        "total_tables": len(result["tables"]),
        "external_link_count": len(result["external_links"]),
        "has_vba": result["has_vba"],
        "risk_flags": []
    }

    # Add risk flags
    if result["has_vba"]:
        result["summary"]["risk_flags"].append("Contains VBA macros")
    if result["has_hidden_sheets"]:
        result["summary"]["risk_flags"].append("Has hidden sheets")
    if len(result["external_links"]) > 0:
        result["summary"]["risk_flags"].append(f"External links: {len(result['external_links'])}")
    if any(s["merged_cells_count"] > 50 for s in result["sheets"]):
        result["summary"]["risk_flags"].append("Heavy use of merged cells")

    wb.close()
    return result


def extract_structure(filepath: str) -> dict:
    """Extract structure from Excel file, auto-detecting format.

    Supports:
    - XLSX/XLSM/XLSB (Excel 2007+): Full support via openpyxl
    - XLS (Excel 97-2003): Basic support via xlrd
    """
    path = Path(filepath)
    ext = path.suffix.lower()

    if ext == '.xls':
        if HAS_XLRD:
            return extract_structure_xls(filepath)
        else:
            return {
                "filename": path.name,
                "format": "xls",
                "support_level": "unsupported",
                "error": "xlrd not installed. Install with: pip install xlrd",
                "sheets": [],
                "summary": {"risk_flags": ["XLS format requires xlrd library"]}
            }
    elif ext in ('.xlsx', '.xlsm', '.xlsb'):
        return extract_structure_xlsx(filepath)
    else:
        return {
            "filename": path.name,
            "format": ext,
            "support_level": "unsupported",
            "error": f"Unsupported format: {ext}",
            "sheets": [],
            "summary": {"risk_flags": [f"Unsupported format: {ext}"]}
        }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_structure.py <excel_file>")
        sys.exit(1)

    filepath = sys.argv[1]
    result = extract_structure(filepath)
    print(json.dumps(result, indent=2, default=str))
