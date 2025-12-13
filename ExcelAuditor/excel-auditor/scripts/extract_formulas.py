#!/usr/bin/env python3
"""Extract and analyze formulas from an Excel file for auditing."""

import sys
import json
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Try to import xlrd for XLS support
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# Excel error values to detect
EXCEL_ERRORS = {'#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#GETTING_DATA'}

# Volatile functions that recalculate on every change
VOLATILE_FUNCTIONS = {'NOW', 'TODAY', 'RAND', 'RANDBETWEEN', 'INDIRECT', 'OFFSET', 'INFO', 'CELL'}

# Common function categories
FUNCTION_CATEGORIES = {
    'lookup': {'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'XLOOKUP', 'LOOKUP', 'CHOOSE'},
    'financial': {'NPV', 'IRR', 'XIRR', 'XNPV', 'PMT', 'FV', 'PV', 'RATE', 'NPER', 'SLN', 'DB', 'DDB'},
    'statistical': {'AVERAGE', 'STDEV', 'VAR', 'MEDIAN', 'MODE', 'CORREL', 'FORECAST', 'TREND', 'GROWTH'},
    'logical': {'IF', 'IFS', 'AND', 'OR', 'NOT', 'SWITCH', 'IFERROR', 'IFNA'},
    'aggregation': {'SUM', 'SUMIF', 'SUMIFS', 'COUNT', 'COUNTIF', 'COUNTIFS', 'SUMPRODUCT', 'AGGREGATE'},
    'text': {'CONCATENATE', 'CONCAT', 'LEFT', 'RIGHT', 'MID', 'LEN', 'TRIM', 'SUBSTITUTE', 'TEXT'},
    'date': {'DATE', 'YEAR', 'MONTH', 'DAY', 'EOMONTH', 'EDATE', 'NETWORKDAYS', 'WORKDAY'},
}

def parse_cell_reference(ref: str) -> tuple:
    """Parse a cell reference like 'A1' or '$A$1' into (col, row)."""
    ref = ref.replace('$', '')
    match = re.match(r'^([A-Za-z]+)(\d+)$', ref)
    if match:
        return (match.group(1).upper(), int(match.group(2)))
    return None


def normalize_formula_pattern(formula: str) -> str:
    """Normalize a formula to detect structural patterns.

    Replaces row numbers with {R} and specific values with placeholders
    to enable pattern matching across rows.
    """
    # Replace row numbers (keeping column letters intact)
    pattern = re.sub(r'([A-Z]+)(\d+)', r'\1{R}', formula)
    # Replace literal numbers (but not in function names)
    pattern = re.sub(r'(?<![A-Z])(\d+\.?\d*)(?![A-Z{])', '{N}', pattern)
    # Replace quoted strings
    pattern = re.sub(r'"[^"]*"', '"{S}"', pattern)
    return pattern


def detect_formula_inconsistencies(wb, formulas_by_cell: dict) -> list:
    """Detect inconsistent formulas within columns - THE SMOKING GUN DETECTOR.

    Finds places where a formula pattern is consistent across most of a column
    but breaks at specific rows. This often indicates:
    - Manual overrides hiding problems
    - Copy/paste errors
    - Intentional manipulation

    Returns list of forensic findings with narrative descriptions.
    """
    findings = []

    # Group formulas by sheet and column
    col_formulas = defaultdict(list)  # (sheet, col) -> [(row, formula, cell_addr)]

    for cell_addr, formula in formulas_by_cell.items():
        # Parse cell address
        if '!' in cell_addr:
            sheet, cell_ref = cell_addr.split('!', 1)
        else:
            sheet = 'Sheet1'
            cell_ref = cell_addr

        parsed = parse_cell_reference(cell_ref)
        if parsed:
            col, row = parsed
            col_formulas[(sheet, col)].append((row, formula, cell_addr))

    # Analyze each column for pattern breaks
    for (sheet, col), cell_list in col_formulas.items():
        if len(cell_list) < 3:
            continue  # Need at least 3 cells to detect pattern breaks

        # Sort by row
        cell_list.sort(key=lambda x: x[0])

        # Normalize formulas to patterns
        patterns = [(row, normalize_formula_pattern(formula), formula, addr)
                   for row, formula, addr in cell_list]

        # Count pattern occurrences
        pattern_counts = defaultdict(list)
        for row, pattern, formula, addr in patterns:
            pattern_counts[pattern].append((row, formula, addr))

        # Find dominant pattern (if any)
        if not pattern_counts:
            continue

        sorted_patterns = sorted(pattern_counts.items(), key=lambda x: -len(x[1]))
        dominant_pattern, dominant_cells = sorted_patterns[0]

        if len(dominant_cells) < 3:
            continue  # No clear pattern

        total_cells = len(cell_list)
        dominant_pct = len(dominant_cells) / total_cells

        # Look for pattern breaks (minority patterns)
        if dominant_pct >= 0.7 and len(sorted_patterns) > 1:
            for minority_pattern, minority_cells in sorted_patterns[1:]:
                if len(minority_cells) <= 2:  # Anomaly detected!
                    for row, formula, addr in minority_cells:
                        # Check if this is surrounded by dominant pattern
                        dominant_rows = [r for r, _, _ in dominant_cells]
                        prev_rows = [r for r in dominant_rows if r < row]
                        next_rows = [r for r in dominant_rows if r > row]

                        if prev_rows and next_rows:  # Surrounded by dominant pattern
                            severity = "critical"
                            narrative = f"SUSPICIOUS: Formula pattern break detected"
                        elif prev_rows or next_rows:
                            severity = "high"
                            narrative = f"WARNING: Formula pattern deviation detected"
                        else:
                            continue

                        # Build forensic narrative
                        sample_dominant = dominant_cells[0][1]  # Sample formula from dominant pattern

                        findings.append({
                            "type": "formula_inconsistency",
                            "severity": severity,
                            "cell": addr,
                            "sheet": sheet,
                            "column": col,
                            "row": row,
                            "actual_formula": formula,
                            "expected_pattern": dominant_pattern,
                            "sample_expected": sample_dominant,
                            "pattern_adherence": f"{len(dominant_cells)}/{total_cells} cells follow pattern",
                            "narrative": f"{narrative} in column {col} of '{sheet}'. "
                                        f"Row {row} has '{formula}' while {len(dominant_cells)} other cells "
                                        f"follow pattern '{sample_dominant[:60]}...'. "
                                        f"This could indicate a manual override, copy error, or intentional manipulation."
                        })

    return findings


def detect_hardcoded_overrides(wb, max_rows: int = 2000) -> list:
    """Detect hardcoded values in columns that predominantly contain formulas.

    THE MAGIC NUMBER FINDER - spots where someone replaced a formula with a
    hardcoded value, often to hide calculation problems or manipulate results.

    Returns list of forensic findings.
    """
    findings = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if not ws.max_row or not ws.max_column:
            continue

        # Analyze each column
        for col_idx in range(1, min(ws.max_column + 1, 100)):  # Limit columns
            col_letter = get_column_letter(col_idx)

            formula_cells = []
            value_cells = []

            # Start from row 2 (skip headers)
            for row_idx in range(2, min(ws.max_row + 1, max_rows)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_addr = f"{sheet_name}!{col_letter}{row_idx}"

                if cell.value is None:
                    continue

                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_cells.append((row_idx, cell.value, cell_addr))
                elif isinstance(cell.value, (int, float)):
                    value_cells.append((row_idx, cell.value, cell_addr))

            # Look for columns that are mostly formulas but have some values
            total = len(formula_cells) + len(value_cells)
            if total < 5:
                continue

            formula_pct = len(formula_cells) / total

            # If column is 70%+ formulas, flag hardcoded values as suspicious
            if formula_pct >= 0.7 and value_cells:
                # Get header for context
                header_cell = ws.cell(row=1, column=col_idx)
                header = str(header_cell.value) if header_cell.value else f"Column {col_letter}"

                for row_idx, value, cell_addr in value_cells:
                    # Check if surrounded by formulas
                    formula_rows = [r for r, _, _ in formula_cells]
                    prev_formula = [r for r in formula_rows if r < row_idx]
                    next_formula = [r for r in formula_rows if r > row_idx]

                    if prev_formula and next_formula:
                        severity = "high"
                        narrative_prefix = "SUSPICIOUS OVERRIDE"
                    elif prev_formula or next_formula:
                        severity = "medium"
                        narrative_prefix = "POTENTIAL OVERRIDE"
                    else:
                        continue

                    # Check if value looks like a "magic number"
                    is_round = isinstance(value, (int, float)) and value == round(value, 0)
                    is_large = isinstance(value, (int, float)) and abs(value) > 10000

                    if is_round and is_large:
                        severity = "critical" if severity == "high" else "high"
                        narrative_prefix = "CRITICAL: Suspicious round number"

                    # Sample neighboring formula
                    sample_formula = formula_cells[0][1] if formula_cells else "Unknown"

                    findings.append({
                        "type": "hardcoded_override",
                        "severity": severity,
                        "cell": cell_addr,
                        "sheet": sheet_name,
                        "column": col_letter,
                        "column_header": header,
                        "row": row_idx,
                        "hardcoded_value": value,
                        "column_formula_ratio": f"{len(formula_cells)}/{total} cells have formulas",
                        "sample_formula": sample_formula[:100],
                        "is_round_number": is_round,
                        "narrative": f"{narrative_prefix} in '{header}' (column {col_letter}) of '{sheet_name}'. "
                                    f"Row {row_idx} contains hardcoded value {value:,.2f} while "
                                    f"{len(formula_cells)} other cells in this column contain formulas like "
                                    f"'{sample_formula[:50]}...'. This value may have been manually inserted "
                                    f"to override a calculation."
                    })

    return findings


def detect_hidden_content(wb) -> dict:
    """Build complete inventory of hidden content - sheets, rows, columns.

    Hidden content is a major forensic red flag - often used to conceal
    calculations, data, or manipulation.
    """
    inventory = {
        "hidden_sheets": [],
        "hidden_rows": [],
        "hidden_columns": [],
        "very_hidden_sheets": [],  # VeryHidden is extra suspicious
        "total_hidden_cells_estimate": 0,
        "narrative": []
    }

    # Check sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Check sheet visibility
        if ws.sheet_state == 'hidden':
            inventory["hidden_sheets"].append({
                "name": sheet_name,
                "rows": ws.max_row,
                "cols": ws.max_column
            })
            inventory["narrative"].append(
                f"HIDDEN SHEET: '{sheet_name}' is hidden from view but contains "
                f"{ws.max_row} rows and {ws.max_column} columns of data."
            )
        elif ws.sheet_state == 'veryHidden':
            inventory["very_hidden_sheets"].append({
                "name": sheet_name,
                "rows": ws.max_row,
                "cols": ws.max_column
            })
            inventory["narrative"].append(
                f"VERY HIDDEN SHEET (VBA-hidden): '{sheet_name}' is programmatically hidden "
                f"and cannot be unhidden through normal Excel UI. Contains {ws.max_row} rows. "
                f"This level of concealment is highly suspicious."
            )

        # Check for hidden rows (sample first 1000 rows)
        if ws.max_row:
            for row_idx in range(1, min(ws.max_row + 1, 1000)):
                try:
                    rd = ws.row_dimensions.get(row_idx)
                    if rd and rd.hidden:
                        inventory["hidden_rows"].append({
                            "sheet": sheet_name,
                            "row": row_idx
                        })
                except:
                    pass

        # Check for hidden columns (sample first 50 columns)
        if ws.max_column:
            for col_idx in range(1, min(ws.max_column + 1, 50)):
                try:
                    col_letter = get_column_letter(col_idx)
                    cd = ws.column_dimensions.get(col_letter)
                    if cd and cd.hidden:
                        inventory["hidden_columns"].append({
                            "sheet": sheet_name,
                            "column": col_letter
                        })
                except:
                    pass

    # Estimate hidden cell count
    hidden_cell_count = 0
    for hs in inventory["hidden_sheets"] + inventory["very_hidden_sheets"]:
        hidden_cell_count += (hs.get("rows", 0) or 0) * (hs.get("cols", 0) or 0)
    hidden_cell_count += len(inventory["hidden_rows"]) * 50  # Estimate 50 cols per hidden row
    hidden_cell_count += len(inventory["hidden_columns"]) * 100  # Estimate 100 rows per hidden col
    inventory["total_hidden_cells_estimate"] = hidden_cell_count

    # Add summary narratives
    if inventory["hidden_rows"]:
        sheets_with_hidden = set(h["sheet"] for h in inventory["hidden_rows"])
        inventory["narrative"].append(
            f"HIDDEN ROWS: {len(inventory['hidden_rows'])} hidden rows detected across "
            f"{len(sheets_with_hidden)} sheet(s). Hidden rows can conceal intermediate "
            f"calculations or problematic data."
        )

    if inventory["hidden_columns"]:
        sheets_with_hidden = set(h["sheet"] for h in inventory["hidden_columns"])
        inventory["narrative"].append(
            f"HIDDEN COLUMNS: {len(inventory['hidden_columns'])} hidden columns detected across "
            f"{len(sheets_with_hidden)} sheet(s). Hidden columns often contain source data "
            f"or calculations that were intentionally concealed."
        )

    if not inventory["narrative"]:
        inventory["narrative"].append("No hidden content detected.")

    return inventory


def calculate_risk_score(forensic_findings: dict) -> dict:
    """Calculate cell, sheet, and workbook level risk scores.

    Produces a 0-100 risk score based on all detected issues.
    """
    risk = {
        "workbook_score": 0,
        "sheet_scores": {},
        "high_risk_cells": [],
        "risk_factors": [],
        "risk_level": "low"  # low, medium, high, critical
    }

    score = 0
    factors = []

    # Circular references (critical)
    circular = forensic_findings.get("circular_references", [])
    if circular:
        score += min(30, len(circular) * 10)
        factors.append(f"Circular references: {len(circular)} cells (+{min(30, len(circular) * 10)})")

    # Formula inconsistencies (high)
    inconsistencies = forensic_findings.get("formula_inconsistencies", [])
    critical_inconsistencies = [i for i in inconsistencies if i.get("severity") == "critical"]
    if critical_inconsistencies:
        score += min(25, len(critical_inconsistencies) * 8)
        factors.append(f"Critical formula inconsistencies: {len(critical_inconsistencies)} (+{min(25, len(critical_inconsistencies) * 8)})")
    elif inconsistencies:
        score += min(15, len(inconsistencies) * 5)
        factors.append(f"Formula inconsistencies: {len(inconsistencies)} (+{min(15, len(inconsistencies) * 5)})")

    # Hardcoded overrides (high)
    hardcoded = forensic_findings.get("hardcoded_overrides", [])
    critical_hardcoded = [h for h in hardcoded if h.get("severity") in ("critical", "high")]
    if critical_hardcoded:
        score += min(20, len(critical_hardcoded) * 6)
        factors.append(f"Suspicious hardcoded values: {len(critical_hardcoded)} (+{min(20, len(critical_hardcoded) * 6)})")

    # Hidden content (medium to high)
    hidden = forensic_findings.get("hidden_content", {})
    very_hidden = len(hidden.get("very_hidden_sheets", []))
    if very_hidden:
        score += very_hidden * 15
        factors.append(f"VeryHidden sheets (VBA-concealed): {very_hidden} (+{very_hidden * 15})")

    hidden_sheets = len(hidden.get("hidden_sheets", []))
    if hidden_sheets:
        score += hidden_sheets * 5
        factors.append(f"Hidden sheets: {hidden_sheets} (+{hidden_sheets * 5})")

    hidden_rows = len(hidden.get("hidden_rows", []))
    if hidden_rows > 5:
        score += min(10, hidden_rows // 2)
        factors.append(f"Hidden rows: {hidden_rows} (+{min(10, hidden_rows // 2)})")

    # Excel errors (medium)
    errors = forensic_findings.get("errors_found", [])
    if errors:
        score += min(10, len(errors) * 2)
        factors.append(f"Excel errors: {len(errors)} (+{min(10, len(errors) * 2)})")

    # Volatile functions (low)
    volatile = forensic_findings.get("volatile_functions", [])
    if len(volatile) > 10:
        score += 5
        factors.append(f"Excessive volatile functions: {len(volatile)} (+5)")

    # VBA macros (medium - depends on context)
    if forensic_findings.get("has_vba"):
        score += 10
        factors.append("Contains VBA macros (+10)")

    # Normalize to 0-100
    score = min(100, score)

    # Determine risk level
    if score >= 60:
        risk_level = "critical"
    elif score >= 40:
        risk_level = "high"
    elif score >= 20:
        risk_level = "medium"
    else:
        risk_level = "low"

    risk["workbook_score"] = score
    risk["risk_factors"] = factors
    risk["risk_level"] = risk_level

    # Identify high-risk cells
    high_risk_cells = []
    for inc in inconsistencies:
        if inc.get("severity") in ("critical", "high"):
            high_risk_cells.append({
                "cell": inc["cell"],
                "reason": "Formula pattern break",
                "severity": inc["severity"]
            })
    for hc in hardcoded:
        if hc.get("severity") in ("critical", "high"):
            high_risk_cells.append({
                "cell": hc["cell"],
                "reason": f"Hardcoded override ({hc.get('hardcoded_value', '?')})",
                "severity": hc["severity"]
            })

    risk["high_risk_cells"] = high_risk_cells[:20]  # Limit to top 20

    return risk


def generate_forensic_narrative(forensic_findings: dict, filename: str) -> str:
    """Generate a Sherlock Holmes style narrative summary of findings.

    This is the executive summary for the forensic report.
    """
    risk = forensic_findings.get("risk_assessment", {})
    risk_score = risk.get("workbook_score", 0)
    risk_level = risk.get("risk_level", "low")

    # Opening based on risk level
    if risk_level == "critical":
        opening = f"CRITICAL FINDINGS in '{filename}': This workbook exhibits multiple indicators of potential manipulation or serious structural problems requiring immediate attention."
    elif risk_level == "high":
        opening = f"SIGNIFICANT CONCERNS identified in '{filename}': Analysis reveals patterns consistent with manual overrides and structural inconsistencies that warrant investigation."
    elif risk_level == "medium":
        opening = f"MODERATE RISK detected in '{filename}': Several anomalies were identified that should be reviewed, though they may have legitimate explanations."
    else:
        opening = f"LOW RISK assessment for '{filename}': No significant red flags detected. Structure appears consistent with normal spreadsheet usage."

    sections = [opening, ""]

    # Key findings
    if risk.get("risk_factors"):
        sections.append("KEY RISK FACTORS:")
        for factor in risk["risk_factors"][:5]:
            sections.append(f"  • {factor}")
        sections.append("")

    # Circular references
    circular = forensic_findings.get("circular_references", [])
    if circular:
        sections.append(f"CIRCULAR REFERENCES ({len(circular)} cells):")
        sections.append(f"  Cells {', '.join(circular[:5])} form calculation loops that may produce unstable or incorrect results.")
        sections.append("")

    # Formula inconsistencies
    inconsistencies = forensic_findings.get("formula_inconsistencies", [])
    if inconsistencies:
        sections.append(f"FORMULA PATTERN BREAKS ({len(inconsistencies)} detected):")
        for inc in inconsistencies[:3]:
            sections.append(f"  • {inc['narrative'][:200]}")
        sections.append("")

    # Hardcoded overrides
    hardcoded = forensic_findings.get("hardcoded_overrides", [])
    critical_hc = [h for h in hardcoded if h.get("severity") in ("critical", "high")]
    if critical_hc:
        sections.append(f"SUSPICIOUS HARDCODED VALUES ({len(critical_hc)} high-severity):")
        for hc in critical_hc[:3]:
            sections.append(f"  • {hc['narrative'][:200]}")
        sections.append("")

    # Hidden content
    hidden = forensic_findings.get("hidden_content", {})
    if hidden.get("narrative") and hidden["narrative"][0] != "No hidden content detected.":
        sections.append("HIDDEN CONTENT:")
        for narr in hidden["narrative"][:3]:
            sections.append(f"  • {narr}")
        sections.append("")

    # Closing recommendation
    if risk_level in ("critical", "high"):
        sections.append("RECOMMENDATION: This workbook should be thoroughly reviewed by a qualified analyst. Consider requesting original source data and audit trail documentation.")
    elif risk_level == "medium":
        sections.append("RECOMMENDATION: Review flagged cells for legitimate business explanations. Document findings for audit trail.")
    else:
        sections.append("RECOMMENDATION: Standard review procedures apply. No urgent action required.")

    return "\n".join(sections)

def extract_references_from_formula(formula: str) -> list:
    """Extract cell and range references from a formula."""
    refs = []
    # Match cell references like A1, $A$1, Sheet1!A1, 'Sheet Name'!A1
    pattern = r"(?:'[^']+'!|[A-Za-z_]+!)?(?:\$?[A-Za-z]+\$?\d+(?::\$?[A-Za-z]+\$?\d+)?)"
    matches = re.findall(pattern, formula)
    return matches

def extract_functions_from_formula(formula: str) -> list:
    """Extract function names from a formula."""
    pattern = r'([A-Z][A-Z0-9_.]+)\s*\('
    return re.findall(pattern, formula.upper())

def calculate_nesting_depth(formula: str) -> int:
    """Calculate maximum nesting depth of a formula."""
    max_depth = 0
    current_depth = 0
    for char in formula:
        if char == '(':
            current_depth += 1
            max_depth = max(max_depth, current_depth)
        elif char == ')':
            current_depth -= 1
    return max_depth

def detect_circular_references(wb, formulas_by_cell: dict) -> list:
    """Detect circular references by building a dependency graph.

    Returns list of cells involved in circular references.
    """
    # Build dependency graph: cell -> cells it depends on
    dependencies = {}

    for cell_addr, formula in formulas_by_cell.items():
        refs = extract_references_from_formula(formula)
        # Normalize cell address
        cell_upper = cell_addr.upper()
        sheet_name = cell_upper.split('!')[0] if '!' in cell_upper else None

        normalized_refs = []
        for ref in refs:
            # Strip $ signs and uppercase for normalization
            ref_clean = ref.replace('$', '').upper()
            if '!' in ref_clean:
                normalized_refs.append(ref_clean)
            elif sheet_name:
                # Add sheet name if missing (also uppercased)
                normalized_refs.append(f"{sheet_name}!{ref_clean}")
            else:
                normalized_refs.append(ref_clean)
        dependencies[cell_upper] = normalized_refs

    # Detect cycles using DFS with proper cycle detection
    circular_cells = set()

    def find_cycle(start_cell, current_cell, path, visited_in_path):
        """DFS to find cycles starting from start_cell."""
        if current_cell in visited_in_path:
            # Found a cycle!
            if current_cell == start_cell or current_cell in path:
                # Mark all cells in the cycle
                try:
                    cycle_idx = path.index(current_cell)
                    for c in path[cycle_idx:]:
                        circular_cells.add(c)
                    circular_cells.add(current_cell)
                except ValueError:
                    circular_cells.add(current_cell)
            return True

        visited_in_path.add(current_cell)
        path.append(current_cell)

        for dep in dependencies.get(current_cell, []):
            # Handle range references - check first cell
            if ':' in dep:
                dep = dep.split(':')[0]

            # Try to find this dep in our dependencies
            if dep in dependencies:
                if find_cycle(start_cell, dep, path, visited_in_path):
                    circular_cells.add(current_cell)

        path.pop()
        visited_in_path.remove(current_cell)
        return False

    # Check from each cell
    for cell in dependencies:
        find_cycle(cell, cell, [], set())

    return list(circular_cells)


def infer_purpose_detailed(function_usage: dict, sheet_names: list, headers: list,
                           formula_count: int, named_range_count: int) -> dict:
    """Infer workbook purpose from multiple signals.

    Returns dict with purpose, confidence, and reasoning.
    """
    signals = []
    scores = defaultdict(float)

    # Analyze function usage patterns
    financial_funcs = {'NPV', 'IRR', 'XIRR', 'XNPV', 'PMT', 'FV', 'PV', 'RATE', 'NPER', 'SLN', 'DB', 'DDB'}
    lookup_funcs = {'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'XLOOKUP', 'LOOKUP', 'CHOOSE'}
    aggregation_funcs = {'SUM', 'SUMIF', 'SUMIFS', 'COUNT', 'COUNTIF', 'COUNTIFS', 'SUMPRODUCT', 'AVERAGE'}
    date_funcs = {'DATE', 'YEAR', 'MONTH', 'DAY', 'EOMONTH', 'EDATE', 'NETWORKDAYS', 'WORKDAY', 'NOW', 'TODAY'}
    text_funcs = {'CONCATENATE', 'CONCAT', 'LEFT', 'RIGHT', 'MID', 'LEN', 'TRIM', 'SUBSTITUTE', 'TEXT'}

    financial_count = sum(function_usage.get(f, 0) for f in financial_funcs)
    lookup_count = sum(function_usage.get(f, 0) for f in lookup_funcs)
    aggregation_count = sum(function_usage.get(f, 0) for f in aggregation_funcs)
    date_count = sum(function_usage.get(f, 0) for f in date_funcs)

    if financial_count > 3:
        scores['financial_model'] += financial_count * 2
        signals.append(f"Financial functions: {financial_count}")

    if lookup_count > 5:
        scores['data_lookup'] += lookup_count * 1.5
        signals.append(f"Lookup functions: {lookup_count}")

    if aggregation_count > 10:
        scores['reporting'] += aggregation_count
        signals.append(f"Aggregation functions: {aggregation_count}")

    if date_count > 5:
        scores['scheduling'] += date_count * 1.5
        signals.append(f"Date functions: {date_count}")

    # Analyze sheet names
    sheet_names_lower = [s.lower() for s in sheet_names]

    financial_keywords = ['income', 'balance', 'cashflow', 'cash flow', 'pl', 'p&l', 'revenue',
                         'expense', 'budget', 'forecast', 'dcf', 'valuation', 'model']
    reporting_keywords = ['summary', 'report', 'dashboard', 'overview', 'analysis', 'kpi']
    data_keywords = ['data', 'raw', 'input', 'source', 'lookup', 'reference', 'master', 'list']
    schedule_keywords = ['schedule', 'timeline', 'calendar', 'gantt', 'plan', 'project']

    for sheet in sheet_names_lower:
        for kw in financial_keywords:
            if kw in sheet:
                scores['financial_model'] += 5
                signals.append(f"Sheet '{sheet}' matches financial pattern")
        for kw in reporting_keywords:
            if kw in sheet:
                scores['reporting'] += 5
                signals.append(f"Sheet '{sheet}' matches reporting pattern")
        for kw in data_keywords:
            if kw in sheet:
                scores['data_lookup'] += 3
        for kw in schedule_keywords:
            if kw in sheet:
                scores['scheduling'] += 5
                signals.append(f"Sheet '{sheet}' matches scheduling pattern")

    # 3-statement model detection
    has_income = any('income' in s or 'p&l' in s or 'pl' in s for s in sheet_names_lower)
    has_balance = any('balance' in s or 'bs' in s for s in sheet_names_lower)
    has_cashflow = any('cash' in s or 'cf' in s for s in sheet_names_lower)

    if sum([has_income, has_balance, has_cashflow]) >= 2:
        scores['financial_model'] += 20
        signals.append("Detected 3-statement model structure")

    # Analyze headers for patterns
    headers_lower = [h.lower() for h in headers if h]

    if any('revenue' in h or 'sales' in h for h in headers_lower):
        scores['financial_model'] += 3
    if any('date' in h or 'month' in h or 'year' in h for h in headers_lower):
        scores['scheduling'] += 2
        scores['reporting'] += 2
    if any('id' in h or 'code' in h or 'name' in h for h in headers_lower):
        scores['data_lookup'] += 2

    # Structural signals
    if named_range_count > 10:
        scores['financial_model'] += 5
        signals.append(f"High named range count: {named_range_count}")

    if formula_count > 500:
        scores['financial_model'] += 5
        scores['reporting'] += 3
        signals.append(f"High formula count: {formula_count}")
    elif formula_count < 10:
        scores['data_storage'] += 10
        signals.append("Low formula count - likely data storage")

    # Determine winner
    if not scores:
        return {
            "purpose": "general_spreadsheet",
            "confidence": 0.3,
            "reasoning": ["No strong signals detected"],
            "all_scores": {}
        }

    sorted_scores = sorted(scores.items(), key=lambda x: -x[1])
    winner = sorted_scores[0]
    runner_up = sorted_scores[1] if len(sorted_scores) > 1 else (None, 0)

    # Calculate confidence
    total = sum(scores.values())
    confidence = winner[1] / total if total > 0 else 0

    # Boost confidence if clear winner
    if runner_up[1] > 0 and winner[1] / runner_up[1] > 2:
        confidence = min(0.95, confidence + 0.2)

    return {
        "purpose": winner[0],
        "confidence": round(confidence, 2),
        "reasoning": signals[:5],
        "all_scores": dict(sorted_scores[:4])
    }


def extract_formulas(filepath: str, max_cells: int = 50000) -> dict:
    """Extract and analyze all formulas from Excel file."""
    result = {
        "filename": Path(filepath).name,
        "formulas": [],
        "errors_found": [],
        "volatile_functions": [],
        "function_usage": defaultdict(int),
        "complexity_metrics": {},
        "issues": [],
        "formula_patterns": defaultdict(list),  # Track similar formulas
        "circular_references": [],  # NEW: Track circular refs
        "purpose_analysis": {},  # NEW: Detailed purpose inference
    }

    try:
        wb = load_workbook(filepath, read_only=False, data_only=False)
    except Exception as e:
        result["error"] = f"Failed to load: {str(e)}"
        return result

    cells_processed = 0
    formulas_by_pattern = defaultdict(list)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = []

        for row in ws.iter_rows():
            if cells_processed >= max_cells:
                result["truncated"] = True
                break

            for cell in row:
                cells_processed += 1

                # Check for formula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    cell_addr = f"{sheet_name}!{cell.coordinate}"

                    formula_info = {
                        "cell": cell_addr,
                        "formula": formula[:500],  # Truncate very long formulas
                        "length": len(formula),
                        "functions": extract_functions_from_formula(formula),
                        "references": extract_references_from_formula(formula)[:20],
                        "nesting_depth": calculate_nesting_depth(formula),
                    }

                    # Track function usage
                    for func in formula_info["functions"]:
                        result["function_usage"][func] += 1

                    # Check for volatile functions
                    volatile_used = [f for f in formula_info["functions"] if f in VOLATILE_FUNCTIONS]
                    if volatile_used:
                        result["volatile_functions"].append({
                            "cell": cell_addr,
                            "volatile_functions": volatile_used
                        })

                    # Flag complex formulas
                    if formula_info["nesting_depth"] > 3:
                        result["issues"].append({
                            "type": "high_nesting",
                            "severity": "warning",
                            "cell": cell_addr,
                            "detail": f"Nesting depth: {formula_info['nesting_depth']}"
                        })

                    if len(formula) > 200:
                        result["issues"].append({
                            "type": "long_formula",
                            "severity": "info",
                            "cell": cell_addr,
                            "detail": f"Formula length: {len(formula)} chars"
                        })

                    # Create pattern key for grouping similar formulas
                    pattern_key = re.sub(r'\d+', '#', formula)[:100]
                    formulas_by_pattern[pattern_key].append(cell_addr)

                    result["formulas"].append(formula_info)

                # Check for error values in calculated results
                elif cell.value in EXCEL_ERRORS:
                    result["errors_found"].append({
                        "cell": f"{sheet_name}!{cell.coordinate}",
                        "error": str(cell.value),
                        "severity": "critical"
                    })

    # Analyze formula consistency
    for pattern, cells in formulas_by_pattern.items():
        if len(cells) >= 3:  # Pattern appears in 3+ cells
            result["formula_patterns"][pattern[:50]] = {
                "count": len(cells),
                "sample_cells": cells[:5]
            }

    # Calculate complexity metrics
    all_nesting = [f["nesting_depth"] for f in result["formulas"]]
    all_lengths = [f["length"] for f in result["formulas"]]

    result["complexity_metrics"] = {
        "total_formulas": len(result["formulas"]),
        "total_errors": len(result["errors_found"]),
        "volatile_function_count": len(result["volatile_functions"]),
        "avg_nesting_depth": round(sum(all_nesting) / len(all_nesting), 2) if all_nesting else 0,
        "max_nesting_depth": max(all_nesting) if all_nesting else 0,
        "avg_formula_length": round(sum(all_lengths) / len(all_lengths), 1) if all_lengths else 0,
        "max_formula_length": max(all_lengths) if all_lengths else 0,
        "unique_patterns": len(formulas_by_pattern),
    }

    # Categorize function usage
    category_counts = defaultdict(int)
    for func, count in result["function_usage"].items():
        for category, funcs in FUNCTION_CATEGORIES.items():
            if func in funcs:
                category_counts[category] += count
                break
    result["function_categories"] = dict(category_counts)

    # Build formulas_by_cell for circular reference detection
    formulas_by_cell = {}
    for f in result["formulas"]:
        formulas_by_cell[f["cell"]] = f["formula"]

    # Detect circular references
    circular_refs = detect_circular_references(wb, formulas_by_cell)
    result["circular_references"] = circular_refs
    if circular_refs:
        result["issues"].append({
            "type": "circular_reference",
            "severity": "critical",
            "cells": circular_refs[:10],
            "detail": f"Found {len(circular_refs)} cells involved in circular references"
        })

    # Collect headers from all sheets for purpose inference
    all_headers = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row and ws.max_row > 0:
            for col in range(1, min(ws.max_column + 1 if ws.max_column else 1, 26)):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    all_headers.append(str(cell.value))

    # Detailed purpose inference
    result["purpose_analysis"] = infer_purpose_detailed(
        function_usage=dict(result["function_usage"]),
        sheet_names=wb.sheetnames,
        headers=all_headers,
        formula_count=len(result["formulas"]),
        named_range_count=len(wb.defined_names) if hasattr(wb, 'defined_names') else 0
    )

    # Keep old field for backwards compatibility but use new analysis
    result["inferred_purpose"] = result["purpose_analysis"]["purpose"]

    # === FORENSIC ANALYSIS ===

    # Detect formula inconsistencies (THE SMOKING GUN DETECTOR)
    result["formula_inconsistencies"] = detect_formula_inconsistencies(wb, formulas_by_cell)
    for inc in result["formula_inconsistencies"]:
        result["issues"].append({
            "type": "formula_inconsistency",
            "severity": inc["severity"],
            "cell": inc["cell"],
            "detail": inc["narrative"][:200]
        })

    # Detect hardcoded overrides (THE MAGIC NUMBER FINDER)
    result["hardcoded_overrides"] = detect_hardcoded_overrides(wb)
    for hc in result["hardcoded_overrides"]:
        if hc["severity"] in ("critical", "high"):
            result["issues"].append({
                "type": "hardcoded_override",
                "severity": hc["severity"],
                "cell": hc["cell"],
                "detail": hc["narrative"][:200]
            })

    # Inventory hidden content
    result["hidden_content"] = detect_hidden_content(wb)

    # Calculate risk score
    forensic_data = {
        "circular_references": result["circular_references"],
        "formula_inconsistencies": result["formula_inconsistencies"],
        "hardcoded_overrides": result["hardcoded_overrides"],
        "hidden_content": result["hidden_content"],
        "errors_found": result["errors_found"],
        "volatile_functions": result["volatile_functions"],
        "has_vba": False  # Will be set by structure extraction
    }
    result["risk_assessment"] = calculate_risk_score(forensic_data)

    # Generate forensic narrative (include risk_assessment in forensic_data)
    forensic_data["risk_assessment"] = result["risk_assessment"]
    result["forensic_narrative"] = generate_forensic_narrative(forensic_data, result["filename"])

    # Convert defaultdicts to regular dicts for JSON
    result["function_usage"] = dict(result["function_usage"])
    result["formula_patterns"] = dict(result["formula_patterns"])

    wb.close()
    return result

def extract_formulas_xls(filepath: str, max_cells: int = 50000) -> dict:
    """Extract formulas from XLS (Excel 97-2003) files using xlrd.

    Note: xlrd has limited formula support - it can detect formula cells
    but getting the actual formula text requires formatting_info=True
    which may not work on all files.
    """
    result = {
        "filename": Path(filepath).name,
        "format": "xls",
        "support_level": "basic",
        "formulas": [],
        "errors_found": [],
        "volatile_functions": [],
        "function_usage": defaultdict(int),
        "complexity_metrics": {},
        "issues": [],
        "formula_patterns": defaultdict(list),
        "circular_references": [],
        "purpose_analysis": {},
    }

    try:
        # Try to load with formatting info for formula access
        wb = xlrd.open_workbook(filepath, formatting_info=True, on_demand=False)
    except Exception:
        try:
            # Fall back to basic loading
            wb = xlrd.open_workbook(filepath, on_demand=True)
        except Exception as e:
            result["error"] = f"Failed to load XLS: {str(e)}"
            return result

    cells_processed = 0
    all_headers = []
    formulas_by_cell = {}

    for idx, sheet_name in enumerate(wb.sheet_names()):
        try:
            ws = wb.sheet_by_index(idx)

            # Collect headers
            if ws.nrows > 0:
                for col in range(min(ws.ncols, 25)):
                    try:
                        val = ws.cell_value(0, col)
                        if val:
                            all_headers.append(str(val))
                    except:
                        pass

            # Scan for formulas and errors
            for row in range(min(ws.nrows, 1000)):  # Limit rows for performance
                if cells_processed >= max_cells:
                    result["truncated"] = True
                    break

                for col in range(ws.ncols):
                    cells_processed += 1
                    try:
                        cell = ws.cell(row, col)
                        cell_addr = f"{sheet_name}!{xlrd.colname(col)}{row + 1}"

                        # Check for formula cells (ctype 2 with XL_CELL_FORMULA would be ideal)
                        # xlrd's cell.ctype: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error, 6=blank
                        if cell.ctype == xlrd.XL_CELL_ERROR:
                            error_code = cell.value
                            error_map = {
                                0x00: '#NULL!', 0x07: '#DIV/0!', 0x0F: '#VALUE!',
                                0x17: '#REF!', 0x1D: '#NAME?', 0x24: '#NUM!', 0x2A: '#N/A'
                            }
                            error_str = error_map.get(error_code, f'#ERROR({error_code})')
                            result["errors_found"].append({
                                "cell": cell_addr,
                                "error": error_str,
                                "severity": "critical"
                            })

                        # For XLS, we can't easily get formula text without deeper parsing
                        # But we can check for text that looks like a formula
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            val = str(cell.value)
                            if val.startswith('='):
                                formula = val
                                formula_info = {
                                    "cell": cell_addr,
                                    "formula": formula[:500],
                                    "length": len(formula),
                                    "functions": extract_functions_from_formula(formula),
                                    "references": extract_references_from_formula(formula)[:20],
                                    "nesting_depth": calculate_nesting_depth(formula),
                                }
                                result["formulas"].append(formula_info)
                                formulas_by_cell[cell_addr] = formula

                                # Track function usage
                                for func in formula_info["functions"]:
                                    result["function_usage"][func] += 1

                                # Check for volatile functions
                                volatile_used = [f for f in formula_info["functions"] if f in VOLATILE_FUNCTIONS]
                                if volatile_used:
                                    result["volatile_functions"].append({
                                        "cell": cell_addr,
                                        "volatile_functions": volatile_used
                                    })

                    except Exception:
                        pass

        except Exception as e:
            result["issues"].append({
                "type": "sheet_error",
                "severity": "warning",
                "detail": f"Error processing sheet {sheet_name}: {str(e)[:100]}"
            })

    # Calculate complexity metrics
    all_nesting = [f["nesting_depth"] for f in result["formulas"]]
    all_lengths = [f["length"] for f in result["formulas"]]

    result["complexity_metrics"] = {
        "total_formulas": len(result["formulas"]),
        "total_errors": len(result["errors_found"]),
        "volatile_function_count": len(result["volatile_functions"]),
        "avg_nesting_depth": round(sum(all_nesting) / len(all_nesting), 2) if all_nesting else 0,
        "max_nesting_depth": max(all_nesting) if all_nesting else 0,
        "avg_formula_length": round(sum(all_lengths) / len(all_lengths), 1) if all_lengths else 0,
        "max_formula_length": max(all_lengths) if all_lengths else 0,
        "format_note": "XLS format - limited formula extraction (formula text not always available)"
    }

    # Categorize function usage
    category_counts = defaultdict(int)
    for func, count in result["function_usage"].items():
        for category, funcs in FUNCTION_CATEGORIES.items():
            if func in funcs:
                category_counts[category] += count
                break
    result["function_categories"] = dict(category_counts)

    # Detect circular references if we found formulas
    if formulas_by_cell:
        circular_refs = detect_circular_references(None, formulas_by_cell)
        result["circular_references"] = circular_refs
        if circular_refs:
            result["issues"].append({
                "type": "circular_reference",
                "severity": "critical",
                "cells": circular_refs[:10],
                "detail": f"Found {len(circular_refs)} cells involved in circular references"
            })

    # Purpose inference
    result["purpose_analysis"] = infer_purpose_detailed(
        function_usage=dict(result["function_usage"]),
        sheet_names=wb.sheet_names(),
        headers=all_headers,
        formula_count=len(result["formulas"]),
        named_range_count=len(wb.name_obj_list) if hasattr(wb, 'name_obj_list') else 0
    )
    result["inferred_purpose"] = result["purpose_analysis"]["purpose"]

    # Convert defaultdicts
    result["function_usage"] = dict(result["function_usage"])
    result["formula_patterns"] = dict(result["formula_patterns"])

    wb.release_resources()
    return result


def extract_formulas_dispatch(filepath: str, max_cells: int = 50000) -> dict:
    """Extract formulas from Excel file, auto-detecting format.

    Supports:
    - XLSX/XLSM/XLSB (Excel 2007+): Full support via openpyxl
    - XLS (Excel 97-2003): Basic support via xlrd
    """
    path = Path(filepath)
    ext = path.suffix.lower()

    if ext == '.xls':
        if HAS_XLRD:
            return extract_formulas_xls(filepath, max_cells)
        else:
            return {
                "filename": path.name,
                "format": "xls",
                "support_level": "unsupported",
                "error": "xlrd not installed. Install with: pip install xlrd",
                "formulas": [],
                "complexity_metrics": {},
                "purpose_analysis": {"purpose": "unknown", "confidence": 0}
            }
    elif ext in ('.xlsx', '.xlsm', '.xlsb'):
        result = extract_formulas(filepath, max_cells)
        result["format"] = ext.lstrip('.')
        result["support_level"] = "full"
        return result
    else:
        return {
            "filename": path.name,
            "format": ext,
            "support_level": "unsupported",
            "error": f"Unsupported format: {ext}",
            "formulas": [],
            "complexity_metrics": {},
            "purpose_analysis": {"purpose": "unknown", "confidence": 0}
        }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_formulas.py <excel_file>")
        sys.exit(1)

    filepath = sys.argv[1]
    result = extract_formulas_dispatch(filepath)
    print(json.dumps(result, indent=2, default=str))
